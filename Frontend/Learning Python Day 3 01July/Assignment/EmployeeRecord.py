import datetime
import re
import openpyxl
from fpdf import FPDF

class Employee:
    def __init__(self, name, dob, phone, email):
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email

    def calculate_age(self):
        today = datetime.date.today()
        dob = datetime.datetime.strptime(self.dob, "%d-%m-%Y").date()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age


def validateDate(dob):
    try:
        datetime.datetime.strptime(dob, '%d-%m-%Y')
    except ValueError:
        print("Incorrect data format, should be DD-MM-YYYY")
        return False
    return True
def validateEmail(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    if not re.match(pattern, email):
        print("Please enter a valid email address.")
        return False
    return True

print("Please Enter following details for an Employee :  ")
while True:
    name = input("Enter Employee name: ")
    if not name.isalpha():
        print("Name should contain only alphabets. Please enter valid name.")
    else:
        break
while True:
    dob = input("Enter Employee date of birth: (DD-MM-YYYY) ")
    if not validateDate(dob):
        print("Please enter valid date of birth.")
    else:
        break

while True:
    phone = input("Enter Employee phone number: ")
    if not phone.isdigit():
        print("Phone number should contain only digits. Please enter valid phone number.")
    elif len(phone)!=10:
        print("Phone number should contain 10 digits. Please enter valid phone number.")
    else:
        break

while True:
    email = input("Enter Employee email: ")
    if not validateEmail(email):
        print("Please enter valid email.")
    else:
        break

emp = Employee(name, dob, phone, email)
age = str(emp.calculate_age())
print("Employee Details : ")
print("******************************")
print("Name          : ",emp.name+
    "\nAge           : ",age+
    "\nDate of Birth : ",emp.dob+
    "\nPhone         : ",emp.phone+
    "\nEmail         : ",emp.email)

print("******************************")


print("Please Enter following details for an Employee :  ")
while True:
    choice = input("Select where you want to save the employee details:\n1) PDF\n2) Excel\n3) txt\nEnter your choice: ")
    if choice not in ['1', '2', '3']:
        print("Invalid choice. Please enter a valid option.")
    else:
        break

if choice == '1':
    # Save employee details as PDF
    try: 
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=13)

        pdf.cell(200, 10, txt="Employee Details:", ln=True, align="C")
        pdf.cell(200, 10, txt="Name: " + emp.name, ln=True, align="L")
        pdf.cell(200, 10, txt="Age: " + age, ln=True, align="L")
        pdf.cell(200, 10, txt="Date of Birth: " + emp.dob, ln=True, align="L")
        pdf.cell(200, 10, txt="Phone: " + emp.phone, ln=True, align="L")
        pdf.cell(200, 10, txt="Email: " + emp.email, ln=True, align="L")

        if pdf.y > 250:
            pdf.add_page()

        pdf.output("employee_details.pdf", "a")
        print("Employee details saved as PDF.")
    except Exception as e:
        print("Error occurred while saving employee details as PDF:", str(e))
elif choice == '2':
    try:
        workbook = openpyxl.load_workbook('employee_details.xlsx')
        sheet = workbook.active
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = emp.name
        sheet.cell(row=next_row, column=2).value = age
        sheet.cell(row=next_row, column=3).value = emp.dob
        sheet.cell(row=next_row, column=4).value = emp.phone
        sheet.cell(row=next_row, column=5).value = emp.email

        workbook.save('employee_details.xlsx')
        print("Employee details saved as Excel.")
    except Exception as e:
        print("Error occurred while saving employee details as Excel:", str(e))
elif choice == '3':
    with open('employee_details.txt', 'a') as file:
            file.write("\n******************************\n")
            file.write("Employee Details:\n")
            file.write("Name:          "+emp.name)
            file.write("\nAge:           "+age)
            file.write("\nDate of Birth: "+emp.dob)
            file.write("\nPhone:         "+emp.phone)
            file.write("\nEmail:         "+emp.email)
    print("Employee details saved as txt.")

